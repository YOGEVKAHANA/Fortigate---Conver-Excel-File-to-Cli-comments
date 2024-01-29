import openpyxl
from tkinter import Tk, filedialog

def open_file_dialog():
    root = Tk()
    root.withdraw()  # Hide the main window

    file_path = filedialog.askopenfilename(
        title="Select Excel File",
        filetypes=[("Excel files", "*.xlsx;*.xls")]
    )

    return file_path

def save_file_dialog():
    root = Tk()
    root.withdraw()  # Hide the main window

    file_path = filedialog.asksaveasfilename(
        title="Save the File",
        defaultextension=".txt",
        filetypes=[("Text files", "*.txt")]
    )

    return file_path

def process_excel_file(excel_file_path, notepad_file_path):
    wb = openpyxl.load_workbook(excel_file_path)

    notepad_content = ""

    # Iterate through the sheets in the workbook
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]

        # Building configuration commands for the current sheet
        sheet_commands = ""
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if sheet_name == "interface_name":  # Replace with the actual name of the first sheet
                interface_name, vdom, mode, ip, security_mode, access = row

                # Building configuration commands for each row
                config_commands = f"""
config system interfaces
edit {interface_name}
    set vdom {vdom}
    set mode {mode}
    set ip {ip}
    set security_mode {security_mode}
    set access {access}
next
"""

            elif sheet_name == "policy_id":  # Replace with the actual name of the second sheet
                policy_id, name, srcaddr, srcintf, dstintf, dstaddr, service, action = row

                # Building configuration commands for each row
                config_commands = f"""
edit {policy_id}
    set name {name}
    set srcintf {srcintf}
    set dstintf {dstintf}
    set srcaddr {srcaddr}
    set dstaddr {dstaddr}
    set service {service}
    set action {action}
next
"""
            elif sheet_name == "Object":  # Replace with the actual name of the third sheet
                name, subnet, comments = row

                # Building configuration commands for each row
                config_commands = f"""
edit {name}
    set subnet {subnet}
    set comments {comments}
next
"""
            elif sheet_name == "vlan": # Replace with the actual name of the third sheet
                vlan, vdom, interface, type, vlanid, mode, ip, allowaccess = row

                # Building configuration commands
                config_commands = f"""
                edit {vlan}
                    set vdom {vdom}
                    set interface  {interface}
                    set type  {type}
                    set vlanid  {vlanid}
                    set mode  {mode}
                    set ip  {ip}
                    set allowaccess  {allowaccess}


                next
                """

            sheet_commands += config_commands

        # Adding the sheet commands to notepad_content
        notepad_content += f"----- Sheet: {sheet_name} -----\n{sheet_commands}"

    # Adding "end" at the end of notepad_content
    notepad_content += "end"

    # Write the notepad_content to a text file
    with open(notepad_file_path, 'w') as notepad_file:
        notepad_file.write(notepad_content)

    print(f"Configuration commands written to {notepad_file_path}")

if __name__ == "__main__":
    excel_file_path = open_file_dialog()

    if excel_file_path:
        notepad_file_path = save_file_dialog()

        if notepad_file_path:
            process_excel_file(excel_file_path, notepad_file_path)
        else:
            print("No location selected to save the file.")
    else:
        print("No file selected.")
